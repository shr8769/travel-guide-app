import tkinter as tk
from tkinter import *
from tkinter import messagebox, ttk, filedialog
from tkinter.scrolledtext import ScrolledText
import wikipedia as wiki
import requests
import webbrowser
from googletrans import Translator, LANGUAGES
from datetime import datetime
import pytz
import pandas as pd
import openpyxl
from opencage.geocoder import OpenCageGeocode
import tkintermapview


# Global variables
main_window = None
login_window = None

# --- Login System (Appears First) ---
def start_login():
    global login_window
    login_window = Tk()
    login_window.title('Travel App - Login Required')
    login_window.geometry('400x350')
    login_window.configure(bg='#57a1f8')
    login_window.resizable(False, False)
    
    # Center the login window
    login_window.update_idletasks()
    x = (login_window.winfo_screenwidth() // 2) - (400 // 2)
    y = (login_window.winfo_screenheight() // 2) - (350 // 2)
    login_window.geometry(f'400x350+{x}+{y}')

    def submit_login():
        user = entry_username.get()
        password = entry_password.get()
        
        if not user or not password or user == 'Username' or password == 'Password':
            error_label.config(text="Please enter username and password")
            return
            
        # Check predefined credentials first
        if user == "admin" and password == "admin123":
            messagebox.showinfo("Login Success", f"Welcome {user}!")
            login_window.destroy()
            create_main_window()
            return
            
        # Check custom accounts from Excel
        try:
            wb = openpyxl.load_workbook('Book1.xlsx')
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                if user == row[0] and password == row[1]:
                    messagebox.showinfo("Login Success", f"Welcome {user}!")
                    login_window.destroy()
                    create_main_window()
                    return
            error_label.config(text="Invalid username or password")
        except FileNotFoundError:
            error_label.config(text="Invalid username or password")

    def on_enter(event, entry, default_text):
        if entry.get() == default_text:
            entry.delete(0, END)
            if default_text == 'Password':
                entry.config(show='*')

    def on_leave(event, entry, default_text):
        if entry.get() == '':
            entry.insert(0, default_text)
            if default_text == 'Password':
                entry.config(show='')

    def open_signup():
        create_account()

    # Login Window UI
    title_label = Label(login_window, text="Travel App Login", fg="white", 
                       font=("Microsoft Yahei UI Light", 24, "bold"), bg="#57a1f8")
    title_label.pack(pady=30)

    info_label = Label(login_window, text="Default: admin / admin123", fg="lightgray", 
                      font=("Microsoft Yahei UI Light", 10), bg="#57a1f8")
    info_label.pack()

    entry_username = Entry(login_window, font=("Microsoft Yahei UI Light", 12), 
                          bg="white", fg="#57a1f8", width=25, bd=0)
    entry_username.pack(pady=10)
    entry_username.insert(0, 'Username')
    entry_username.bind('<FocusIn>', lambda event: on_enter(event, entry_username, 'Username'))
    entry_username.bind('<FocusOut>', lambda event: on_leave(event, entry_username, 'Username'))

    entry_password = Entry(login_window, font=("Microsoft Yahei UI Light", 12), 
                          bg="white", fg="#57a1f8", width=25, bd=0)
    entry_password.pack(pady=10)
    entry_password.insert(0, 'Password')
    entry_password.bind('<FocusIn>', lambda event: on_enter(event, entry_password, 'Password'))
    entry_password.bind('<FocusOut>', lambda event: on_leave(event, entry_password, 'Password'))

    error_label = Label(login_window, text="", fg="red", bg="#57a1f8", 
                       font=("Microsoft Yahei UI Light", 10))
    error_label.pack(pady=5)

    login_button = Button(login_window, text="Login", 
                         font=("Microsoft Yahei UI Light", 14, "bold"), 
                         bg="white", fg="#57a1f8", bd=0, width=15, 
                         command=submit_login, cursor="hand2")
    login_button.pack(pady=20)

    signup_link = Button(login_window, text="Don't have an account? Create one here", 
                        font=("Microsoft Yahei UI Light", 10), 
                        bg="#57a1f8", fg="white", bd=0, 
                        command=open_signup, cursor="hand2")
    signup_link.pack(pady=10)

    login_window.mainloop()

# --- Account Creation Function ---
def create_account():
    def submit():
        user_ = entry_create_username.get()
        pass_ = entry_create_password.get()
        if not user_ or not pass_ or user_ == 'Username' or pass_ == 'Password':
            error_label.config(text="Please enter valid username and password")
            return
        try:
            wb = openpyxl.load_workbook('Book1.xlsx')
            sheet = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet['A1'] = 'Username'
            sheet['B1'] = 'Password'
        if user_ in [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]:
            error_label.config(text="Username already exists")
        else:
            sheet.append([user_, pass_])
            wb.save('Book1.xlsx')
            messagebox.showinfo("Success", "Account created successfully!\nYou can now login.")
            root_signup.destroy()

    def on_enter(event, entry, default_text):
        if entry.get() == default_text:
            entry.delete(0, END)
            if default_text == 'Password':
                entry.config(show='*')

    def on_leave(event, entry, default_text):
        if entry.get() == '':
            entry.insert(0, default_text)
            if default_text == 'Password':
                entry.config(show='')

    # Determine parent window
    if main_window:
        parent = main_window
    else:
        parent = login_window

    root_signup = Toplevel(parent)
    root_signup.resizable(False, False)
    root_signup.title('Create Account')
    root_signup.geometry('400x350')
    root_signup.configure(bg='#57a1f8')

    title_label = Label(root_signup, text="Create Account", fg="white", 
                       font=("Microsoft Yahei UI Light", 24, "bold"), bg="#57a1f8")
    title_label.pack(pady=30)

    entry_create_username = Entry(root_signup, font=("Microsoft Yahei UI Light", 12), 
                                 bg="white", fg="#57a1f8", width=25, bd=0)
    entry_create_username.pack(pady=10)
    entry_create_username.insert(0, 'Username')
    entry_create_username.bind('<FocusIn>', lambda event: on_enter(event, entry_create_username, 'Username'))
    entry_create_username.bind('<FocusOut>', lambda event: on_leave(event, entry_create_username, 'Username'))

    entry_create_password = Entry(root_signup, font=("Microsoft Yahei UI Light", 12), 
                                 bg="white", fg="#57a1f8", width=25, bd=0)
    entry_create_password.pack(pady=10)
    entry_create_password.insert(0, 'Password')
    entry_create_password.bind('<FocusIn>', lambda event: on_enter(event, entry_create_password, 'Password'))
    entry_create_password.bind('<FocusOut>', lambda event: on_leave(event, entry_create_password, 'Password'))

    error_label = Label(root_signup, text="", fg="red", bg="#57a1f8", 
                       font=("Microsoft Yahei UI Light", 10))
    error_label.pack(pady=5)

    create_button = Button(root_signup, text="Create Account", 
                          font=("Microsoft Yahei UI Light", 14, "bold"), 
                          bg="white", fg="#57a1f8", bd=0, width=15, 
                          command=submit, cursor="hand2")
    create_button.pack(pady=20)

# --- All Feature Functions ---
def init():
    def get_timezone(country):
        country_mapping = {
            "USA": "America/New_York",
            "UK": "Europe/London",
            'India': "Asia/Kolkata",
            'SaudiArabia': "Asia/Riyadh",
        }
        timezone_name = country_mapping.get(country, country)
        return pytz.timezone(timezone_name)
    def get_time():
        country = entry.get()
        try:
            timezone = get_timezone(country)
            time_now = datetime.now(timezone)
            formatted_time = time_now.strftime("%Y-%m-%d %H:%M:%S %Z")
            result_label.config(text=f"Current time in {country}: {formatted_time}")
        except pytz.UnknownTimeZoneError:
            result_label.config(text="Invalid country or timezone")
    root = Toplevel(main_window)
    root.geometry('400x400')
    root.config(bg='#57a1f8')
    root.title("World Clock App")
    label = tk.Label(root, text="Enter country:")
    label.pack(pady=10)
    entry = tk.Entry(root)
    entry.pack(pady=10)
    result_label = tk.Label(root, text="")
    result_label.pack(pady=10)
    submit_button = tk.Button(root, text="Get Time", command=get_time)
    submit_button.pack(pady=10)

def search():
    search_data = search_entry.get()
    try:
        data = wiki.summary(search_data, sentences=5, auto_suggest=False)
        search_entry.delete(0, END)
        text.delete(1.0, END)
        text.insert(END, data)
    except wiki.exceptions.DisambiguationError:
        messagebox.showwarning('Disambiguation', f"Multiple results found for '{search_data}'. Please be more specific.")
    except wiki.exceptions.HTTPTimeoutError:
        messagebox.showerror('Error', 'Error connecting to Wikipedia. Please check your internet connection.')
    except wiki.exceptions.PageError:
        messagebox.showwarning('Page Not Found', f"No results found for '{search_data}'.")
    except Exception:
        messagebox.showerror('Error', 'An unexpected error occurred. Please try again later.')

def open_search_window():
    global search_entry, text
    root_info = Toplevel(main_window)
    root_info.title('Search Application')
    root_info.geometry('400x500')
    root_info.config(bg='white')
    search_entry = Entry(root_info, width=25, font=('Microsoft Yahei UI Light', 12), bd=10, relief=RIDGE)
    search_entry.place(x=15, y=20)
    search_lbl = Label(root_info, text='Searching result for:', font=('Microsoft Yahei UI Light', 12, 'bold'), bg='#57a1f8')
    search_lbl.place(x=15, y=70)
    text = ScrolledText(root_info, font=('Microsoft Yahei UI Light', 10), bd=0, relief=SUNKEN)
    text.place(x=15, y=100, height=400, width=300)
    search_btn = Button(root_info, text='Search', font=('Microsoft Yahei UI Light', 12, 'bold'), width=10, fg='white', bg='#57a1f8', command=search)
    search_btn.place(x=10, y=420)
    clear_btn = Button(root_info, text='Clear', font=('Microsoft Yahei UI Light', 12, 'bold'), width=10, fg='white', bg='#57a1f8', command=lambda: text.delete(1.0, END))
    clear_btn.place(x=105, y=420)
    exit_btn = Button(root_info, text='Exit', font=('arial', 12, 'bold'), width=10, fg='white', bg='#57a1f8', command=root_info.destroy)
    exit_btn.place(x=200, y=420)

def open_weather_window():
    global city_entry, location_label, icon_label, temperature_label, description_label
    root_wea = Toplevel(main_window)
    root_wea.title('Weather Application')
    root_wea.geometry('400x400')
    city_entry = Entry(root_wea, font=('Microsoft Yahei UI Light', 12))
    city_entry.pack(pady=10)
    search_button = Button(root_wea, text='Search', command=search_weather, font=('Microsoft Yahei UI Light', 12, 'bold'), width=10, fg='white', bg='#57a1f8')
    search_button.pack(pady=10)
    location_label = Label(root_wea, font=('Microsoft Yahei UI Light', 12))
    location_label.pack()
    icon_label = Label(root_wea)
    icon_label.pack()
    temperature_label = Label(root_wea, font=('Microsoft Yahei UI Light', 12))
    temperature_label.pack()
    description_label = Label(root_wea, font=('Microsoft Yahei UI Light', 12))
    description_label.pack()

def search_weather():
    city = city_entry.get()
    result = get_weather(city)
    if result is None:
        return
    icon_url, temperature, description, city, country = result
    location_label.configure(text=f"{city}, {country}")
    temperature_label.configure(text=f"Temperature: {temperature:.2f}°C")
    description_label.configure(text=f"Description: {description}")

def get_weather(city):
    API_key = '130f28eb96516e7e58e6f5ecb4cc82d7'  # Replace with your API key
    url = f"https://api.openweathermap.org/data/2.5/weather?q={city}&appid={API_key}"
    res = requests.get(url)
    if res.status_code == 404:
        messagebox.showerror('Error', 'City not found')
        return None
    weather = res.json()
    icon_id = weather['weather'][0]['icon']
    temperature = weather['main']['temp'] - 273.15
    description = weather['weather'][0]['description']
    city = weather['name']
    country = weather['sys']['country']
    icon_url = f"https://openweathermap.org/img/wn/{icon_id}@2x.png"
    return (icon_url, temperature, description, city, country)

def create_popup():
    global popup
    popup = Toplevel(main_window)
    popup.title("Popup Window")
    popup.geometry("300x100")
    label = Label(popup, text="Please use this in EMERGENCY only")
    label.pack(padx=20, pady=20)
    close_button = Button(popup, text="PROCEED", command=ask_question_box)
    close_button.pack(pady=10)

def ask_question_box():
    global popup
    result = messagebox.askquestion("Question", "INITIATE EMERGENCY SOS??")
    if result == 'yes':
        create_popup2()
        if popup:
            popup.destroy()
            popup = None

def create_popup2():
    popup2 = Toplevel(main_window)
    popup2.title("SOS")
    popup2.geometry("600x150")
    Label(popup2, text="EMERGENCY SOS INITIATED!").pack(padx=20, pady=10)
    Label(popup2, text="Your SOS is received. Emergency services are deployed. Stay calm; help is coming.").pack()
    Label(popup2, text="Your safety is our priority").pack()
    Button(popup2, text="OKAY", command=popup2.destroy).pack(pady=10)

def convert(base_entry, target_entry, amount_entry, result_label):
    base_currency = base_entry.get().upper()
    target_currency = target_entry.get().upper()
    amount = amount_entry.get()
    try:
        response = requests.get(f'https://api.exchangerate-api.com/v4/latest/{base_currency}')
        data = response.json()
        exchange_rate = data['rates'][target_currency]
        converted_amount = float(amount) * exchange_rate
        result_label.config(text=f'{amount} {base_currency} = {converted_amount:.2f} {target_currency}')
    except Exception:
        result_label.config(text='Error: Failed to fetch data')

def create_forex_window():
    root_forex = Toplevel(main_window)
    root_forex.title('Currency Converter')
    window_width = 400
    window_height = 300
    screen_width = root_forex.winfo_screenwidth()
    screen_height = root_forex.winfo_screenheight()
    x_coordinate = (screen_width // 2) - (window_width // 2)
    y_coordinate = (screen_height // 2) - (window_height // 2)
    root_forex.geometry(f'{window_width}x{window_height}+{x_coordinate}+{y_coordinate}')
    root_forex.configure(bg='#57a1f8')
    center_frame = Frame(root_forex, bg='#57a1f8')
    center_frame.pack(expand=True)
    base_label = Label(center_frame, text='Base Currency:', bg='#57a1f8', fg='white')
    base_label.grid(row=0, column=0, sticky=E, pady=10)
    base_entry = Entry(center_frame, width=20)
    base_entry.grid(row=0, column=1, sticky=W, pady=10)
    target_label = Label(center_frame, text='Target Currency:', bg='#57a1f8', fg='white')
    target_label.grid(row=1, column=0, sticky=E, pady=10)
    target_entry = Entry(center_frame, width=20)
    target_entry.grid(row=1, column=1, sticky=W, pady=10)
    amount_label = Label(center_frame, text='Amount:', bg='#57a1f8', fg='white')
    amount_label.grid(row=2, column=0, sticky=E, pady=10)
    amount_entry = Entry(center_frame, width=20)
    amount_entry.grid(row=2, column=1, sticky=W, pady=10)
    result_label = Label(center_frame, text='', bg='#57a1f8', fg='white')
    result_label.grid(row=4, columnspan=2)
    convert_button = Button(center_frame, text='Convert', command=lambda: convert(base_entry, target_entry, amount_entry, result_label), bg='white')
    convert_button.grid(row=3, columnspan=2, pady=20)

class MapViewerApp:
    def __init__(self, root, api_key):
        self.root = root
        self.root.title("Map Viewer")
        self.label = Label(root, text="Enter Location:")
        self.label.pack(pady=10)
        self.entry = Entry(root, width=30)
        self.entry.pack(pady=10)
        self.button = Button(root, text="Show Map", command=self.show_map)
        self.button.pack(pady=10)
        self.geocoder = OpenCageGeocode(api_key)
    def show_map(self):
        location = self.entry.get()
        result = self.geocoder.geocode(location)
        if result and len(result):
            lat, lng = result[0]['geometry']['lat'], result[0]['geometry']['lng']
            map_url = f"https://www.google.com/maps/place/{lat},{lng}"
            webbrowser.open(map_url)
        else:
            messagebox.showerror("Error", "Location not found. Please enter a valid location.")


def open_map_viewer():
    root_map = Toplevel()
    root_map.geometry("800x600")
    root_map.title("Tkinter Map Viewer")

    # Entry for address
    entry = Entry(root_map, width=50)
    entry.pack(pady=10)

    # Map widget
    map_widget = tkintermapview.TkinterMapView(root_map, width=780, height=500, corner_radius=0)
    map_widget.pack(pady=10)

    def show_address():
        address = entry.get()
        if address:
            map_widget.set_address(address)
            map_widget.set_zoom(8)  # or any zoom level you prefer

    # Button to show address
    btn = Button(root_map, text="Show Map", command=show_address)
    btn.pack(pady=5)

    # Optionally, set a default location
    map_widget.set_address("Berlin, Germany")
    map_widget.set_zoom(5)

def translate_now():
    textn = text1.get(1.0, END)
    t1 = Translator()
    src_lang = combo1.get()
    dest_lang = combo2.get()
    src_code = [k for k, v in LANGUAGES.items() if v.lower() == src_lang.lower()]
    dest_code = [k for k, v in LANGUAGES.items() if v.lower() == dest_lang.lower()]
    if not src_code or not dest_code:
        text2.delete(1.0, END)
        text2.insert(END, "Invalid language selection.")
        return
    trans_text = t1.translate(textn, src=src_code[0], dest=dest_code[0])
    trans_text = trans_text.text
    text2.delete(1.0, END)
    text2.insert(END, trans_text)

def open_translator_window():
    global root_trans, combo1, combo2, text1, text2
    root_trans = Toplevel(main_window)
    root_trans.title("Simple Translator")
    root_trans.geometry('700x350')
    root_trans.config(bg='#57a1f8')
    language_list = list(LANGUAGES.values())
    combo1 = ttk.Combobox(root_trans, values=language_list, state="readonly")
    combo1.place(x=30, y=20)
    combo1.set("english")
    combo2 = ttk.Combobox(root_trans, values=language_list, state="readonly")
    combo2.place(x=400, y=20)
    combo2.set("select language")
    text1 = ScrolledText(root_trans,  bg="white", relief=GROOVE, wrap=WORD)
    text1.place(x=30, y=60, width=250, height=200)
    text2 = ScrolledText(root_trans,  bg="white", relief=GROOVE, wrap=WORD)
    text2.place(x=400, y=60, width=250, height=200)
    trans_btn = Button(root_trans, text='Translate', command=translate_now, bg='#57a1f8', fg='white', bd=10)
    trans_btn.place(x=300, y=300)

class ItineraryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Itinerary Planner")
        self.root.config(bg='#57a1f8')
        self.days = []
        self.day_label = Label(root, text="Day:")
        self.day_label.grid(row=0, column=0, padx=10, pady=10, sticky="W")
        self.day_var = StringVar()
        self.day_entry = Entry(root, textvariable=self.day_var)
        self.day_entry.grid(row=0, column=1, padx=10, pady=10, sticky="W")
        self.destination_label = Label(root, text="Destination:")
        self.destination_label.grid(row=1, column=0, padx=10, pady=10, sticky="W")
        self.destination_var = StringVar()
        self.destination_entry = Entry(root, textvariable=self.destination_var)
        self.destination_entry.grid(row=1, column=1, padx=10, pady=10, sticky="W")
        self.activities_label = Label(root, text="Activities:")
        self.activities_label.grid(row=2, column=0, padx=10, pady=10, sticky="W")
        self.activities_var = StringVar()
        self.activities_entry = Entry(root, textvariable=self.activities_var)
        self.activities_entry.grid(row=2, column=1, padx=10, pady=10, sticky="W")
        self.time_label = Label(root, text="Time:")
        self.time_label.grid(row=3, column=0, padx=10, pady=10, sticky="W")
        self.time_var = StringVar()
        self.time_entry = Entry(root, textvariable=self.time_var)
        self.time_entry.grid(row=3, column=1, padx=10, pady=10, sticky="W")
        self.add_button = Button(root, text="Add Day", command=self.add_day)
        self.add_button.grid(row=4, column=0, columnspan=2, pady=20)
        self.schedule_tree = ttk.Treeview(root, columns=('Day', 'Destination', 'Activities', 'Time'), show='headings')
        self.schedule_tree.heading('Day', text='Day')
        self.schedule_tree.heading('Destination', text='Destination')
        self.schedule_tree.heading('Activities', text='Activities')
        self.schedule_tree.heading('Time', text='Time')
        self.schedule_tree.grid(row=5, column=0, columnspan=2, padx=10, pady=10)
        self.export_button = Button(root, text="Export to Excel", command=self.export_to_excel)
        self.export_button.grid(row=6, column=0, columnspan=2, pady=20)
    def add_day(self):
        day = self.day_var.get()
        destination = self.destination_var.get()
        activities = self.activities_var.get()
        time = self.time_var.get()
        if day and destination and activities and time:
            self.days.append((day, destination, activities, time))
            self.update_schedule()
            self.clear_entries()
        else:
            messagebox.showwarning("Incomplete Information", "Please fill in all fields.")
    def update_schedule(self):
        self.schedule_tree.delete(*self.schedule_tree.get_children())
        for day, destination, activities, time in self.days:
            self.schedule_tree.insert('', 'end', values=(day, destination, activities, time))
    def clear_entries(self):
        self.day_var.set('')
        self.destination_var.set('')
        self.activities_var.set('')
        self.time_var.set('')
    def export_to_excel(self):
        if self.days:
            df = pd.DataFrame(self.days, columns=['Day', 'Destination', 'Activities', 'Time'])
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if file_path:
                df.to_excel(file_path, index=False)
                messagebox.showinfo("Export Successful", f"Data exported to {file_path}")
        else:
            messagebox.showwarning("No Data", "No itinerary data to export.")

def create_itinerary_app():
    itinerary_root = Toplevel(main_window)
    itinerary_root.title('Itinerary Planner')
    app = ItineraryApp(itinerary_root)

class TravelGuideApp:
    def __init__(self, master):
        self.master = master
        self.greetings = ["Hi", "Hola", "Bonjour", "ನಮಸ್ತೆ", "أهلاً"]
        self.guides = [
            "Welcome to your Travel Guide",
            "Bienvenido a tu Guía de Viaje",
            "Bienvenue dans votre guide de voyage",
            "ನಿಮ್ಮ ಪ್ರಯಾಣ ಮಾರ್ಗದರ್ಶಿಗೆ ಸ್ವಾಗತ",
            "مرحبا بكم في دليل السفر الخاص بك"
        ]
        self.label = Label(self.master, text="", font=("Microsoft Yahei UI Light", 12), bg='#f0f0f0')
        self.label.pack(pady=10)
        self.update_label()
    def update_label(self):
        current_greeting = self.greetings.pop(0)
        self.greetings.append(current_greeting)
        current_guide = self.guides.pop(0)
        self.guides.append(current_guide)
        self.label.config(text=f"{current_greeting}! {current_guide}")
        self.master.after(2000, self.update_label)

# --- Main Window (Only appears after login) ---
def create_main_window():
    global main_window
    main_window = Tk()
    main_window.title('Travel Application Hub - Welcome!')
    window_height = 600
    window_width = 700
    screen_width = main_window.winfo_screenwidth()
    screen_height = main_window.winfo_screenheight()
    x_position = (screen_width - window_width) // 2
    y_position = (screen_height - window_height) // 2
    main_window.geometry(f'{window_width}x{window_height}+{x_position}+{y_position}')
    main_window.configure(bg='#f0f0f0')

    header_label = Label(main_window, text="🌍 Travel Application Hub 🌍", 
                        font=("Microsoft Yahei UI Light", 20, "bold"), 
                        bg='#f0f0f0', fg='#57a1f8')
    header_label.pack(pady=20)

    # Account section
    account_frame = Frame(main_window, bg='#f0f0f0')
    account_frame.pack(pady=10)

    logout_btn = Button(account_frame, text='Logout', command=logout, 
                       font=('Microsoft Yahei UI Light', 10, 'bold'), 
                       width=10, fg='white', bg='#ff6b6b', cursor="hand2")
    logout_btn.pack(side=LEFT, padx=5)

    signup_btn = Button(account_frame, text='Create New Account', command=create_account, 
                       font=('Microsoft Yahei UI Light', 10, 'bold'), 
                       width=15, fg='white', bg='#57a1f8', cursor="hand2")
    signup_btn.pack(side=LEFT, padx=5)

    features_label = Label(main_window, text="Available Features", 
                          font=("Microsoft Yahei UI Light", 16, "bold"), 
                          bg='#f0f0f0', fg='#333')
    features_label.pack(pady=(20, 10))

    button_frame = Frame(main_window, bg='#f0f0f0')
    button_frame.pack(pady=10)

    # Feature buttons in a grid layout
    search_btn = Button(button_frame, text='📚 Search', command=open_search_window, 
                       font=('Microsoft Yahei UI Light', 10, 'bold'), 
                       width=15, fg='white', bg='#57a1f8', cursor="hand2")
    search_btn.grid(row=0, column=0, padx=5, pady=5)

    weather_btn = Button(button_frame, text='🌤️ Weather', command=open_weather_window, 
                        font=('Microsoft Yahei UI Light', 10, 'bold'), 
                        width=15, fg='white', bg='#57a1f8', cursor="hand2")
    weather_btn.grid(row=0, column=1, padx=5, pady=5)

    forex_btn = Button(button_frame, text='💱 Currency', command=create_forex_window, 
                      font=('Microsoft Yahei UI Light', 10, 'bold'), 
                      width=15, fg='white', bg='#57a1f8', cursor="hand2")
    forex_btn.grid(row=0, column=2, padx=5, pady=5)

    map_btn = Button(button_frame, text='🗺️ Map Viewer', command=open_map_viewer, 
                    font=('Microsoft Yahei UI Light', 10, 'bold'), 
                    width=15, fg='white', bg='#57a1f8', cursor="hand2")
    map_btn.grid(row=1, column=0, padx=5, pady=5)

    trans_btn = Button(button_frame, text='🔤 Translator', command=open_translator_window, 
                      font=('Microsoft Yahei UI Light', 10, 'bold'), 
                      width=15, fg='white', bg='#57a1f8', cursor="hand2")
    trans_btn.grid(row=1, column=1, padx=5, pady=5)

    timezone_btn = Button(button_frame, text='🕐 Timezone', command=init, 
                         font=('Microsoft Yahei UI Light', 10, 'bold'), 
                         width=15, fg='white', bg='#57a1f8', cursor="hand2")
    timezone_btn.grid(row=1, column=2, padx=5, pady=5)

    itinerary_btn = Button(button_frame, text='📋 Itinerary', command=create_itinerary_app, 
                          font=('Microsoft Yahei UI Light', 10, 'bold'), 
                          width=15, fg='white', bg='#57a1f8', cursor="hand2")
    itinerary_btn.grid(row=2, column=0, padx=5, pady=5)

    sos_btn = Button(button_frame, text='🚨 Emergency SOS', command=create_popup, 
                    font=('Microsoft Yahei UI Light', 10, 'bold'), 
                    width=15, fg='white', bg='red', cursor="hand2")
    sos_btn.grid(row=2, column=1, padx=5, pady=5)

    # Travel Guide (animated text)
    TravelGuideApp(main_window)
    
    main_window.mainloop()

def logout():
    result = messagebox.askquestion("Logout", "Are you sure you want to logout?")
    if result == 'yes':
        main_window.destroy()
        start_login()

# --- Start Application ---
if __name__ == "__main__":
    start_login()
